import os
import re
import sys
import time
import openpyxl
import pyttsx3
import sounddevice as sd
from scipy.io.wavfile import write
import whisper
from word2number import w2n

# Initialize Text-to-Speech
engine = pyttsx3.init()
engine.setProperty('rate', 170)

def speak(text):
    print(f"🔊 Mimi: {text}")
    engine.say(text)
    engine.runAndWait()

# Load tiny.en model (75MB for instant processing)
print("⚡ Loading fast voice engine (tiny.en)...")
whisper_model = whisper.load_model("tiny.en")

EXCEL_FILE = "my_budget.xlsx"

def record_audio(filename, duration, samplerate=16000):
    audio_data = sd.rec(int(duration * samplerate), samplerate=samplerate, channels=1, dtype='int16')
    sd.wait()
    write(filename, samplerate, audio_data)
    return filename

def transcribe_audio(filename):
    result = whisper_model.transcribe(filename, language="en", fp16=False)
    return result['text'].strip()

def update_excel_smart(filename, command):
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(filename)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    cmd_lower = command.lower()

    # 1. Add Column Command
    if "add column" in cmd_lower or "create column" in cmd_lower:
        col_name = re.sub(r'.*(add|create)\s+column\s+(of\s+)?', '', cmd_lower).strip().strip('.')
        col_name = col_name.replace("named", "").replace("name", "Name").capitalize()
        
        max_col = ws.max_column
        target_col = 1 if ws.cell(row=1, column=1).value is None else max_col + 1

        ws.cell(row=1, column=target_col, value=col_name)
        wb.save(filename)
        speak(f"Added column {col_name}")
        return True

    # 2. Extract Cell Coordinates (e.g. B2, C3)
    cell_match = re.search(r'\b([a-zA-Z])\s*(\d+)\b', command)
    if cell_match:
        col_letter = cell_match.group(1).upper()
        row_num = int(cell_match.group(2))
        cell_ref = f"{col_letter}{row_num}"
        
        val_match = re.search(r'\b\d+\b', command)
        val = int(val_match.group(0)) if val_match else command.split()[-1]
        
        ws[cell_ref] = val
        wb.save(filename)
        speak(f"Updated cell {cell_ref} to {val}")
        return True

    # 3. Put Value under Column
    under_match = re.search(r'(put|add|set)\s+(.*?)\s+(under|in|into)\s+(.*)', cmd_lower)
    if under_match:
        val_str = under_match.group(2).strip()
        target_header = under_match.group(4).strip().strip('.')

        try:
            val = w2n.word_to_num(val_str)
        except ValueError:
            val = val_str if not val_str.isdigit() else int(val_str)

        target_col_idx = None
        for col in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col).value
            if header_val and str(header_val).strip().lower() == target_header:
                target_col_idx = col
                break

        if not target_col_idx:
            target_col_idx = 1 if ws.cell(row=1, column=1).value is None else ws.max_column + 1
            ws.cell(row=1, column=target_col_idx, value=target_header.capitalize())

        next_row = 2
        while ws.cell(row=next_row, column=target_col_idx).value is not None:
            next_row += 1

        ws.cell(row=next_row, column=target_col_idx, value=val)
        wb.save(filename)
        speak(f"Added {val} under column {target_header.capitalize()}")
        return True

    speak("I couldn't match that command.")
    return False

def start_mimi():
    speak("Mimi is active!")
    print("\n👂 Listening for 'Hey Mimi'...")

    wake_words = ["mimi", "hey mimi", "hi mimi", "me me", "meemee", "mini"]

    while True:
        # Fast 2-second check for wake word
        audio_file = record_audio("wake_temp.wav", duration=2.0)
        text = transcribe_audio(audio_file).lower()

        if any(wake in text for wake in wake_words):
            speak("How can I help you?")
            
            # Pause 1 second so speaker finishes before mic records
            time.sleep(0.8)

            print("🎙️ SPEAK YOUR COMMAND NOW...")
            cmd_audio = record_audio("command_temp.wav", duration=5.0)
            command_text = transcribe_audio(cmd_audio)
            print(f"You said: '{command_text}'")

            if command_text:
                success = update_excel_smart(EXCEL_FILE, command_text)
                if success:
                    speak("All done! Goodbye.")
                    sys.exit(0)  # Close program
            
            speak("Closing now.")
            sys.exit(0)

if __name__ == "__main__":
    start_mimi()